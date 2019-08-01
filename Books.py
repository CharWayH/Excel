# encoding=utf-8
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image
import datetime
import calendar


# 函数部分
# 数字转周几
def weekdayName(dayName):
    if dayName == 0:
        dayName = '周一'
    elif dayName == 1:
        dayName = '周二'
    elif dayName == 2:
        dayName = '周三'
    elif dayName == 3:
        dayName = '周四'
    elif dayName == 4:
        dayName = '周五'
    elif dayName == 5:
        dayName = '周六'
    else:
        dayName = '周日'
    return dayName


'''
year =datetime.datetime.now().strftime("%Y")
month = datetime.datetime.now().strftime("%m") 
days=calendar.monthrange(int(year),int(month))    #该月共几天
'''

print('请输入年：')
year = int(input())
print('请输入月：')
month = int(input())
days = calendar.monthrange(year,month)

wb = Workbook()
ws= wb.active


# 填内容
# 表单名
ws['B1'] = '%s年%s月住宿登记表'%(year,month)
# 金浦花园
ws['A2'] = '金浦花园'
ws['A4'] = '纽约'
ws['A8'] = '巴黎'
ws['A11'] = '伦敦'
ws['A13'] = '米兰'
ws['A14'] = '厅'
# 床位号
ws['B4'] = '四人间1号床'
ws['B5'] = '四人间2号床'
ws['B6'] = '四人间3号床'
ws['B7'] = '四人间4号床'
ws['B8'] = '三人间1号床'
ws['B9'] = '三人间2号床'
ws['B10'] = '三人间3号床'
ws['B11'] = '两人间1号床'
ws['B12'] = '两人间2号床'
ws['B13'] = '单人间'
ws['B14'] = '沙发床'
ws['B15'] = 'T1'
ws['B16'] = 'T2'


#备注
ws['C17'] = '以上表格是金浦花园小区电梯房8楼，兰色区域是四人间一床75/天，紫色区域是三人间一床80/天，绿色区域是伦敦两人间一床/70/天，粉色区域是米兰单人间一床100/天，白色区域是厅两人间' \
            '一床70/天，另外煮饭费是每人每天3元，要煮饭的自觉发煮饭费给管理员否则发现了按三倍收取煮饭费'
ws['C19'] = '请大家按照预订的床位号入住，从上往下四人间1234号床，三人间123号床，两人间12号床，米兰单人间，厅沙发床，厅单1床，厅单2床'


# 周几及日期
for i in range(1, days[1]+1):
    cal = calendar.weekday(year, month, i)    # 该月份第一天是周几
    dayName = weekdayName(cal)
    ws.cell(row=2, column=2+i).value = dayName    # 周几
    ws.cell(row=3, column=2+i).value = i      # 日期
    ws.cell(row=2, column=2+i).alignment = Alignment(horizontal='center', vertical='center')    # 字体居中
    ws.cell(row=3, column=2+i).alignment = Alignment(horizontal='center', vertical='center')

column_max = i+2  # 最后一天，单元格的列数
# 共用/字体
ws['A3'] = '房间名'
ws['B2'] = '日期'
ws['B3'] = '床号'
ws['A2'].font = Font(color='ff8c69', bold=True, size=14)    # 橘色加粗字体
ws['B1'].font = Font(color='3883c2', size=16, bold=True)    # 红色字体
ws['C17'].font = ws['C19'].font = Font(color='ee3f4d', bold=True)   # 红色加粗字体

for i in range(2, 17):               # B2-B17字体加粗
        ws['B%s' % i].font = Font(bold=True)

bg_column_max = column_max+1    # 背景色最后一行

# 背景颜色
for i in range(4, 8):
    for j in range(3, bg_column_max):
        ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor="f2accf")

for i in range(8, 11):
    for j in range(3, bg_column_max):
        ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor="aff0ee")

for i in range(11, 13):
    for j in range(3, bg_column_max):
        ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor="e7e4b8")

for i in range(14, 17):
    for j in range(3, bg_column_max):
        ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', fgColor="4682b4")

for j in range(3, bg_column_max):
        ws.cell(row=13, column=j).fill = PatternFill(fill_type='solid', fgColor="61ca90")


# 设置单元格格式
# 合并单元格
ws.merge_cells('B1:E1')
ws.merge_cells('A4:A7')
ws.merge_cells('A8:A10')
ws.merge_cells('A11:A12')
ws.merge_cells('A14:A16')
ws.merge_cells('A34:N34')
ws.merge_cells('A35:N35')

# 字体居中
ws['A4'].alignment = ws['A8'].alignment = ws['A11'].alignment = ws['A13'].alignment = ws['A14'].alignment = Alignment(horizontal='center', vertical='center')


for i in range(2, 17):
    for j in range(1,column_max+1):
        ws.cell(row=i, column=j).border = Border(top=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color = 'FF000000'),left = Side(border_style='thin',color = 'FF000000'),bottom = Side(border_style='thin',color = 'FF000000'),)


# 上边加粗
for i in range(4, column_max):
    ws.cell(row=2, column=i).border = ws.cell(row=4, column=i).border = ws.cell(row=8, column=i).border = ws.cell(row=11, column=i).border = ws.cell(row=13, column=i).border = ws.cell(row=14, column=i).border = Border(left=Side(border_style='thin', color='FF000000'), top=Side(border_style='medium', color='FF000000'), right=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
# 上，下边加粗
for i in range(4, column_max):
    Border(left=Side(border_style='thin', color='FF000000'),top=Side(border_style='medium', color='FF000000'), right=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
# 上，左，右加粗
ws.cell(row=14, column=2).border = ws.cell(row=14, column=1).border = ws.cell(row=13, column=2).border = ws.cell(row=13, column=1).border = ws.cell(row=11, column=2).border = ws.cell(row=11, column=1).border = ws.cell(row=8, column=1).border = ws.cell(row=8, column=2).border = ws.cell(row=4, column=2).border = ws.cell(row=2, column=2).border = ws.cell(row=2, column=1).border = Border(left=Side(border_style='medium', color='FF000000'),top=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
# 上，左，下加粗
ws.cell(row=13, column=2).border = ws.cell(row=13, column=1).border = ws.cell(row=11, column=2).border = ws.cell(row=11, column=1).border = ws.cell(row=8, column=1).border = ws.cell(row=2, column=2).border = ws.cell(row=2, column=1).border = Border(left=Side(border_style='medium', color='FF000000'),top=Side(border_style='medium', color='FF000000'), right=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
# 上，左加粗
ws.cell(row=14, column=3).border= ws.cell(row=13, column=3).border = ws.cell(row=11, column=3).border = ws.cell(row=8, column=3).border = ws.cell(row=4, column=3).border = ws.cell(row=2, column=3).border = Border(left=Side(border_style='medium', color='FF000000'),top=Side(border_style='medium', color='FF000000'), right=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
# 右，上加粗
ws.cell(row=4, column=column_max).border = ws.cell(row=14, column=column_max).border = ws.cell(row=2, column=column_max).border = ws.cell(row=11, column=column_max).border = ws.cell(row=8, column=column_max).border = Border(left=Side(border_style='thin', color='FF000000'),top=Side(border_style='medium', color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
# 右，下加粗
ws.cell(row=7, column=column_max).border = ws.cell(row=10, column=column_max).border = ws.cell(row=16, column=column_max).border = ws.cell(row=12, column=column_max).border = ws.cell(row=3, column=column_max).border = ws.cell(row=3, column=column_max).border = Border(left=Side(border_style='thin', color='FF000000'),top=Side(border_style='thin', color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
# 左，下，右加粗
ws.cell(row=16, column=1).border = ws.cell(row=16, column=2).border = ws.cell(row=12, column=2).border = ws.cell(row=10, column=2).border = ws.cell(row=7, column=2).border = ws.cell(row=3, column=2).border = Border(left=Side(border_style='medium', color='FF000000'),top=Side(border_style='thin', color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
# 左，右加粗
ws.cell(row=15, column=2).border = ws.cell(row=9, column=2).border = ws.cell(row=5, column=2).border = ws.cell(row=6, column=2).border = Border(left=Side(border_style = 'medium', color='FF000000'),top=Side(border_style='thin', color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
# 右加粗
ws.cell(row=15, column=column_max).border = ws.cell(row=9, column=column_max).border = ws.cell(row=6, column=column_max).border = ws.cell(row=5, column=column_max).border = Border(left=Side(border_style='thin', color='FF000000'),top=Side(border_style='thin', color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
# 左，下加粗
ws.cell(row=16, column=3).border = Border(left=Side(border_style='medium', color='FF000000'), top=Side(border_style='thin',
color='FF000000'), right=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))
# 上，下，右加粗
ws.cell(row=13, column=column_max).border = Border(left=Side(border_style='thin', color='FF000000'), top=Side(border_style='medium',
color='FF000000'), right=Side(border_style='medium', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))


# 右，下加粗
ws.cell(row=16, column=column_max).border = Border(left=Side(border_style='thin', color='FF000000'),top=Side(border_style='thin', color='FF000000'), right=Side(border_style='medium', color='FF000000'),
                                                   bottom=Side(border_style='medium', color='FF000000'))
# 下边加粗
for i in range(4, column_max):
    ws.cell(row=16, column=i).border = Border(left=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin',
    color='FF000000'), right=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='medium', color='FF000000'))


# 添加图片
img = Image('photo.png')
ws.add_image(img, 'C21')

# 调整单元格大小
ws.column_dimensions['A'].width = 12.0
ws.column_dimensions['B'].width = 13.67

filename = '住宿%s月份.xlsx'%month
wb.save(filename=filename)
