#encoding=utf-8
'''
openpyxlExcel写入的操作
'''
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
wb = Workbook()
ws1 = wb.active #打开表单
ws1.merge_cells('B1:E1')#合并B1到E1间的单元格
ws1['A1']='插入内容'#在A1位置插入内容
ws1['A2'] = datetime.datetime.now().strftime("%Y-%m-%d")#在A2位置插入时间
ws1.append([1, 2, 3])#在下一行分别按列顺序插入1,2,3
ws1.title = '工作表标题1'    #工作表标题
ws2 = wb.create_sheet(title='工作表标题2') #创建一个新的工作表
#循环插入
ws3 = wb.create_sheet(title="Data1")
for row in range(1, 10):    #1-9行
    for col in range(2, 30):    #2-29列
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))   #使用openpyxl.utils工具包
print(ws3['B9'].value)#工作表3,B9位置的值

filename = 'Excel.xlsx' #文件名
wb.save(filename=filename)#保存文件

