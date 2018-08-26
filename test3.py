from openpyxl.reader.excel import load_workbook

wb = load_workbook(filename='电影.xlsx')

#print("Worksheet range(s):",wb.get_named_ranges())
# print("Worksheet name(s):", wb.get_sheet_names())
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
print ("Work Sheet Titile:", ws.title)
print ("Work Sheet Rows:", ws.max_row)
print ("Work Sheet Cols:", ws.max_column)